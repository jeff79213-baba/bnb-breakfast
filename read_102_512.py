#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
早餐統計 102~512 可視匯出
規則:
- 只取「早餐統計」工作表
- 欄位: 房號 | 來源 | 住宿狀態 | 蛋奶素 | 全素 | 大人 | 小孩 | 用餐時間
- 102~512 房號範圍, 依房號排序
- 排除隱藏列 (此檔隱藏列 83~102), 無隱藏欄
- 不 | 加 | 購 → 合併為 不加購
- 續住 佔用蛋奶素 → 抽到 住宿狀態 獨立欄
- Booking/CTrip/Agoda 預設 不加購, 加購=付費
執行: 用主 .venv 的 python 執行
  C:/Users/TW-10/Documents/firebase雲端資料夾/.venv/Scripts/python.exe read_102_512.py
"""
import os, re
import openpyxl

def get_excel_path():
    # 與此 py 同目錄的 02住客記錄.xlsx
    base = os.path.dirname(os.path.abspath(__file__))
    for f in os.listdir(base):
        if "住客" in f and f.endswith(".xlsx") and not f.startswith("~$"):
            return os.path.join(base, f)
    raise FileNotFoundError("找不到 02住客記錄.xlsx")

def read_102_512(excel_path=None, sheet_name="早餐統計"):
    excel_path = excel_path or get_excel_path()
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    hidden_rows = {r for r in ws.row_dimensions if ws.row_dimensions[r].hidden}
    # 找標題列
    hr = None
    for r in ws.iter_rows(min_row=1, max_row=5):
        for c in r:
            if c.value and "房號" in str(c.value):
                hr = c.row
                break
        if hr: break
    if not hr:
        raise ValueError("找不到 房號 標題列")

    rows = []
    for row in ws.iter_rows(min_row=hr+1, max_row=ws.max_row):
        rnum = row[0].row
        if rnum in hidden_rows:
            continue
        b = row[1]  # B 房號
        if b.value is None:
            continue
        room_raw = str(b.value).strip()
        if not re.match(r"^\d+$", room_raw):
            continue
        rn = int(room_raw)
        if not (102 <= rn <= 512):
            continue

        c_val = row[2].value if len(row) > 2 else None
        pD = row[3].value if len(row) > 3 else None
        pE = row[4].value if len(row) > 4 else None
        pF = row[5].value if len(row) > 5 else None
        pG = row[6].value if len(row) > 6 else None
        pH = row[7].value if len(row) > 7 else None

        def s(v): return "" if v is None else str(v).strip()
        pD, pE, pF, pG, pH = s(pD), s(pE), s(pF), s(pG), s(pH)
        src = s(c_val)

        住宿狀態 = 蛋奶素 = 全素 = 大人 = 小孩 = ""
        用餐時間 = pH

        is_bu = (pE == "不" and pF == "加" and pG == "購")
        is_xu = (pD == "續住" and is_bu)

        if is_xu:
            住宿狀態 = "續住"
            全素 = "不加購"
        elif is_bu:
            if pD == "續住":
                住宿狀態 = "續住"
            elif pD:
                蛋奶素 = pD
            全素 = "不加購"
        else:
            if pD == "續住":
                住宿狀態 = "續住"
                全素 = pE
                大人 = pF
                小孩 = pG
            elif pD == "加購":
                蛋奶素 = "加購"
                全素 = pE
                大人 = pF
                小孩 = pG
            else:
                蛋奶素 = pD
                全素 = pE
                大人 = pF
                小孩 = pG

        rows.append((rn, [room_raw, src, 住宿狀態, 蛋奶素, 全素, 大人, 小孩, 用餐時間]))

    rows.sort(key=lambda x: x[0])
    return rows

def main():
    rows = read_102_512()
    head = ["房號","來源","住宿狀態","蛋奶素","全素","大人","小孩","用餐時間"]
    print("\t".join(head))
    for _, vals in rows:
        print("\t".join(vals))
    # 同時存檔到同目錄
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "早餐102-512可視匯出.txt")
    with open(out, "w", encoding="utf-8") as f:
        f.write("\t".join(head)+"\n")
        for _, vals in rows:
            f.write("\t".join(vals)+"\n")
    print(f"\n已輸出 {len(rows)} 筆 -> {out}")

if __name__ == "__main__":
    main()
